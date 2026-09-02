#!/usr/bin/env python3
"""Turn the training-plan JSON export into a review workbook.

Reads the output of scripts/export-training-plans.ts and writes an .xlsx with a
sheet per level of the sync pipeline, so a plan can be followed from the stored
program day all the way down to the individual Garmin workout step.

Usage:
    python3 scripts/build-training-plan-workbook.py plans.json training-plans.xlsx
"""

import json
import math
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
NOTE_FONT = Font(name=FONT, size=10, italic=True, color="595959")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")


def write_sheet(wb, title, headers, rows, widths=None, freeze="A2"):
    """Write one table. Returns the sheet so callers can add formulas after."""
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = freeze
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def flatten_steps(steps, repeat_no=None, repeat_value=None):
    """Yield (repeat_no, repeat_value, step) for a step list, descending into repeats.

    Garmin nests interval work inside WorkoutRepeatStep containers. Flattening
    keeps one row per actual step while preserving which repeat block it sits in
    and how many times that block runs.
    """
    for step in steps:
        if step.get("type") == "WorkoutRepeatStep":
            for inner in flatten_steps(
                step.get("steps", []), step.get("stepOrder"), step.get("repeatValue")
            ):
                yield inner
        else:
            yield (repeat_no, repeat_value, step)


def count_steps(workout):
    if not workout:
        return 0
    return sum(
        len(list(flatten_steps(seg.get("steps", []))))
        for seg in workout.get("segments", [])
    )


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "training-plans.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "training-plans.xlsx"

    data = json.load(open(src))
    programs = data["programs"]

    wb = Workbook()
    wb.remove(wb.active)

    # ── Days ─────────────────────────────────────────────────────────────────
    day_rows = []
    session_rows = []
    step_rows = []
    exercise_rows = []
    flag_rows = []

    for p in programs:
        pid, pname = p["id"], p["name"]
        seen_days = Counter(d["day"] for d in p["days"])

        for d in p["days"]:
            day, title = d["day"], d["title"]
            week = math.ceil(day / 7) if day > 0 else 0
            sessions = d["sessions"]
            mapped = [s for s in sessions if s["workout"]]
            sports = sorted({s["workout"]["sport"] for s in mapped})
            cats = sorted({s["category"] for s in sessions})

            day_rows.append([
                pid, pname, p["programType"], day, week,
                ((day - 1) % 7) + 1 if day > 0 else 0,
                title, d["sourceKind"],
                len(d["exercises"]), len(d["runs"]),
                len(sessions), len(mapped), len(sessions) - len(mapped),
                ", ".join(cats), ", ".join(sports),
                sum(count_steps(s["workout"]) for s in mapped),
                "; ".join(d.get("dataFlags", [])),
            ])

            # Anything stored on the day that never reaches the watch is the
            # single most useful thing to see when reviewing the sync.
            has_content = bool(d["exercises"] or d["runs"])
            if has_content and not sessions:
                flag_rows.append([pid, pname, day, title, "No session generated",
                                  "Day has content but workoutToDays() returned nothing."])
            if has_content and sessions and not mapped:
                flag_rows.append([pid, pname, day, title, "Nothing pushed to Garmin",
                                  f"All {len(sessions)} session(s) mapped to null "
                                  f"(category: {', '.join(cats)}). Nothing appears on the watch."])
            for msg in d.get("dataFlags", []):
                flag_rows.append([pid, pname, day, title, "Stored data shape", msg])
            if seen_days[day] > 1:
                flag_rows.append([pid, pname, day, title, "Duplicate day number",
                                  f"Day {day} appears {seen_days[day]} times in this program."])

            for s in sessions:
                w = s["workout"]
                sid = s["sessionIdx"]
                steps_n = count_steps(w)

                session_rows.append([
                    pid, pname, day, week, title, sid, s["dayKey"],
                    s.get("sessionType") or "", s.get("garminSport") or "",
                    s["category"],
                    "yes" if w else "no",
                    w["workoutName"] if w else "",
                    w["sport"] if w else "",
                    w.get("description", "") if w else "",
                    w.get("estimatedDurationInSecs") if w else None,
                    steps_n,
                ])

                if not w:
                    continue

                # Run fidelity checks. mapRunEasy() never reads the planned
                # distance — it takes a duration from the first "N minutes" in
                # the description text and otherwise falls back to 30 minutes.
                # Both failure modes are invisible until you compare the two.
                if w["sport"] == "RUNNING" and d["runs"]:
                    flat = [st for _, _, st in flatten_steps(
                        [x for seg in w["segments"] for x in seg.get("steps", [])])]
                    planned_km = sum(r.get("distance") or 0 for r in d["runs"])
                    secs = w.get("estimatedDurationInSecs") or 0

                    if s["category"] != "run_intervals":
                        if len(flat) == 1 and secs == 1800:
                            detail = ("No duration in the text, so the step defaulted to 30:00 "
                                      f"— the planned {planned_km}km is never used."
                                      if planned_km > 0 else
                                      "No duration in the text and no planned distance, so the step "
                                      "defaulted to a flat 30:00.")
                            flag_rows.append([pid, pname, day, title, "Run duration defaulted", detail])
                        elif len(flat) == 1 and planned_km > 0 and secs > 0:
                            pace = secs / planned_km
                            if pace < 150 or pace > 900:
                                flag_rows.append([pid, pname, day, title, "Duration / distance mismatch",
                                                  f"Step is {secs // 60}:{secs % 60:02d} for a planned "
                                                  f"{planned_km}km ({pace / 60:.1f} min/km). Duration was "
                                                  f"read from the description text, not the distance."])

                    interval_runs = [r for r in d["runs"]
                                     if (r.get("noIntervals") or 0) > 1
                                     or r.get("type") in ("intervals", "tempo")]
                    if interval_runs and s["category"] != "run_intervals":
                        flag_rows.append([pid, pname, day, title, "Intervals flattened",
                                          f"{len(interval_runs)} interval/tempo run(s) stored, but the day "
                                          f"classified as '{s['category']}' so it maps to one flat easy run. "
                                          f"The interval structure never reaches the watch."])

                for seg in w["segments"]:
                    for repeat_no, repeat_value, st in flatten_steps(seg.get("steps", [])):
                        step_rows.append([
                            pid, pname, day, title, sid, w["workoutName"], seg["sport"],
                            repeat_no or "", repeat_value or "",
                            st["stepOrder"], st["intensity"], st.get("description") or "",
                            st["durationType"], st.get("durationValue"),
                            st["targetType"], st.get("targetValueLow"), st.get("targetValueHigh"),
                            st.get("exerciseCategory") or "", st.get("exerciseName") or "",
                            st.get("weightValue"), st.get("weightDisplayUnit") or "",
                        ])

                # A strength session whose steps carry no exercise identity shows
                # on the watch as unnamed generic work.
                unmatched = [
                    st for _, _, st in flatten_steps(
                        [x for seg in w["segments"] for x in seg.get("steps", [])])
                    if st["intensity"] == "ACTIVE" and not st.get("exerciseCategory")
                ]
                if w["sport"] == "STRENGTH_TRAINING" and unmatched:
                    flag_rows.append([pid, pname, day, title, "Unmatched exercises",
                                      f"{len(unmatched)} of {steps_n} steps have no Garmin "
                                      f"exerciseCategory — they show as generic steps."])

                open_only = all(
                    st["durationType"] == "OPEN"
                    for _, _, st in flatten_steps(
                        [x for seg in w["segments"] for x in seg.get("steps", [])])
                )
                if open_only and steps_n > 1:
                    flag_rows.append([pid, pname, day, title, "No measurable targets",
                                      "Every step is OPEN duration — the watch cannot "
                                      "auto-advance or measure this session."])

            # ── Source rows: what the plan actually stores ───────────────────
            for i, ex in enumerate(d["exercises"]):
                exercise_rows.append([
                    pid, pname, day, week, title, "exercise", i + 1,
                    ex.get("name", ""), ex.get("details", ""),
                    ex.get("sessionType", ""), ex.get("garminSport", ""),
                    ex.get("garminExerciseCategory", ""), ex.get("garminExerciseName", ""),
                    ex.get("sets"), ex.get("reps"), ex.get("weightKg"), ex.get("restSeconds"),
                    "", "", "", "", "",
                ])
            for i, r in enumerate(d["runs"]):
                exercise_rows.append([
                    pid, pname, day, week, title, "run", i + 1,
                    r.get("type", ""), r.get("description", ""),
                    "run", "RUNNING", "", "", "", "", "", "",
                    r.get("distance"), r.get("paceZone", ""), r.get("targetPace"),
                    r.get("effortLevel"), r.get("noIntervals"),
                ])

    # ── Programs (rollups as live formulas over the Days sheet) ──────────────
    prog_rows = []
    for p in programs:
        days = p["days"]
        prog_rows.append([
            p["id"], p["collection"], p["name"], p["programType"],
            p.get("targetRace") or "", p.get("visibility") or "",
            p["assignedUserCount"], p["retainedUserCount"],
            len(days),
            max((d["day"] for d in days), default=0),
            None, None, None, None,  # formulas, filled below
            (p["description"] or "")[:500],
        ])

    ws_prog = write_sheet(
        wb, "Programs",
        ["Program ID", "Collection", "Name", "Type", "Target race", "Visibility",
         "Assigned users", "Retained users", "Days stored", "Highest day no.",
         "Weeks", "Sessions generated", "Pushed to Garmin", "Skipped", "Description"],
        prog_rows,
        widths=[26, 15, 34, 10, 12, 11, 14, 14, 12, 15, 8, 18, 17, 10, 70],
    )
    last = len(day_rows) + 1
    for r in range(2, len(prog_rows) + 2):
        ws_prog.cell(r, 11, f"=ROUNDUP(J{r}/7,0)")
        ws_prog.cell(r, 12, f"=SUMIFS(Days!$K$2:$K${last},Days!$A$2:$A${last},$A{r})")
        ws_prog.cell(r, 13, f"=SUMIFS(Days!$L$2:$L${last},Days!$A$2:$A${last},$A{r})")
        ws_prog.cell(r, 14, f"=SUMIFS(Days!$M$2:$M${last},Days!$A$2:$A${last},$A{r})")
        for c in range(11, 15):
            ws_prog.cell(r, c).font = BODY_FONT

    write_sheet(
        wb, "Days",
        ["Program ID", "Program", "Type", "Day", "Week", "Day of week", "Title",
         "Stored as", "Exercises", "Runs", "Sessions", "Pushed", "Skipped",
         "Category(s)", "Garmin sport(s)", "Total steps", "Data flags"],
        day_rows,
        widths=[26, 30, 10, 7, 7, 12, 40, 11, 10, 7, 10, 9, 9, 26, 24, 12, 46],
    )

    write_sheet(
        wb, "Garmin Workouts",
        ["Program ID", "Program", "Day", "Week", "Day title", "Session #", "Sync key",
         "Session type", "Requested sport", "Category", "Pushed?", "Garmin workout name",
         "Garmin sport", "Garmin description", "Est. duration (s)", "Steps"],
        session_rows,
        widths=[26, 30, 7, 7, 36, 10, 11, 13, 20, 16, 10, 34, 20, 60, 15, 8],
    )

    write_sheet(
        wb, "Garmin Steps",
        ["Program ID", "Program", "Day", "Day title", "Session #", "Workout name", "Sport",
         "In repeat (step)", "Repeat x", "Step order", "Intensity", "Description",
         "Duration type", "Duration value", "Target type", "Target low", "Target high",
         "Exercise category", "Exercise name", "Weight", "Weight unit"],
        step_rows,
        widths=[26, 26, 7, 30, 10, 30, 20, 15, 10, 11, 12, 44, 14, 14, 13, 12, 12, 20, 26, 9, 12],
    )

    write_sheet(
        wb, "Source Detail",
        ["Program ID", "Program", "Day", "Week", "Day title", "Row kind", "#",
         "Name / run type", "Details / description", "Session type", "Garmin sport",
         "Garmin exercise category", "Garmin exercise name", "Sets", "Reps",
         "Weight (kg)", "Rest (s)", "Distance (km)", "Pace zone", "Target pace (s/km)",
         "Effort (RPE)", "Intervals"],
        exercise_rows,
        widths=[26, 26, 7, 7, 34, 10, 5, 30, 62, 13, 20, 24, 26, 7, 7, 12, 9, 13, 12, 19, 12, 10],
    )

    ws_flags = write_sheet(
        wb, "Review Flags",
        ["Program ID", "Program", "Day", "Day title", "Flag", "Detail"],
        flag_rows,
        widths=[26, 30, 7, 40, 26, 88],
    )
    for row in ws_flags.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.fill = FLAG_FILL

    # ── Read Me (first tab) ──────────────────────────────────────────────────
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["A"].width = 118
    lines = [
        ("Training plans — Garmin sync review", TITLE_FONT),
        ("", BODY_FONT),
        (f"Source: {data['source']}", BODY_FONT),
        (f"Generated: {data['generatedAt']}", BODY_FONT),
        ("", BODY_FONT),
        ("How this was produced", Font(name=FONT, bold=True, size=11)),
        ("Every program day was run through the same code path the live sync uses:", BODY_FONT),
        ("    program.workouts[]  →  workoutToDays()  →  mapWorkoutDay()  →  Garmin workout", BODY_FONT),
        ("(src/lib/garmin/program-adapter.ts and src/lib/garmin/workout-mapper.ts, called directly).", BODY_FONT),
        ("So the payloads below are what POST /api/garmin/sync-plan would push, not a re-implementation.", BODY_FONT),
        ("", BODY_FONT),
        ("Sheets", Font(name=FONT, bold=True, size=11)),
        ("Programs         One row per program, with live rollups over the Days sheet.", BODY_FONT),
        ("Days             One row per stored program day: what it holds, and what it produced.", BODY_FONT),
        ("Garmin Workouts  One row per session the sync would create. 'Pushed? = no' means the mapper", BODY_FONT),
        ("                 returned null and nothing reaches the watch for that session.", BODY_FONT),
        ("Garmin Steps     One row per step inside each workout, repeats flattened and labelled.", BODY_FONT),
        ("Source Detail    The stored exercise and run rows the mapping was derived from.", BODY_FONT),
        ("Review Flags     Days and sessions worth a second look — see below.", BODY_FONT),
        ("", BODY_FONT),
        ("Assumptions and limits", Font(name=FONT, bold=True, size=11)),
        ("• Weights show only what is stored on the exercise (weightKg). At sync time resolveWeightKg()", BODY_FONT),
        ("  can derive a load from the athlete's 1RM personal records, so a real push may carry weights", BODY_FONT),
        ("  this export shows as blank. That is per-athlete and cannot be shown in a per-program extract.", BODY_FONT),
        ("• Sync horizon is ignored. The live sync pushes ~14 days from the athlete's start date; this", BODY_FONT),
        ("  export maps every day of every program so the whole plan can be reviewed at once.", BODY_FONT),
        ("• 'Sync key' is the key the sync would write into users/{uid}.garminPlanSync.workouts.", BODY_FONT),
        ("", BODY_FONT),
        ("What the flags mean", Font(name=FONT, bold=True, size=11)),
        ("Nothing pushed to Garmin   The day has content but every session mapped to null. Usually the", BODY_FONT),
        ("                           title-based classifier returned 'skip' — check the day title wording.", BODY_FONT),
        ("Unmatched exercises        Strength steps with no Garmin exerciseCategory; they appear on the", BODY_FONT),
        ("                           watch as unnamed generic steps. Fix by naming the exercise so", BODY_FONT),
        ("                           program-enricher.ts matches it, or by setting the field explicitly.", BODY_FONT),
        ("No measurable targets      Every step is OPEN duration — the watch cannot auto-advance.", BODY_FONT),
        ("Run duration defaulted     mapRunEasy() found no 'N minutes' in the text, so the step is a flat", BODY_FONT),
        ("                           30:00 and the planned distance is ignored.", BODY_FONT),
        ("Duration / distance        The step's duration implies an impossible pace for the planned", BODY_FONT),
        ("mismatch                   distance — the duration was read out of the description text.", BODY_FONT),
        ("Intervals flattened        Interval or tempo runs on a day the classifier did not read as", BODY_FONT),
        ("                           intervals; they map to a single steady run.", BODY_FONT),
        ("Stored data shape          The Firestore document is missing a field the sync code reads.", BODY_FONT),
        ("Duplicate day number       Two stored days share a day number; the later one wins on the watch.", BODY_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(i, 1, text)
        c.font = font
        c.alignment = Alignment(vertical="top")
    ws.sheet_view.showGridLines = False

    wb.save(out)
    print(f"Wrote {out}")
    print(f"  programs {len(prog_rows)} | days {len(day_rows)} | sessions {len(session_rows)} "
          f"| steps {len(step_rows)} | source rows {len(exercise_rows)} | flags {len(flag_rows)}")


if __name__ == "__main__":
    main()
