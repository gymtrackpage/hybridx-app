#!/usr/bin/env python3
"""Build the updated running plan: one tab per program, rebuilt step by step.

Reads the export (current state), the rebuilt run sessions, and the strength
proposals, and writes a workbook that shows what each session should send to
Garmin and how that differs from what it sends today.

Usage:
    python3 scripts/build-updated-plan-workbook.py plans.json rebuilt.json \\
        proposals.json updated-plan.xlsx
"""

import json
import math
import re
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, bold=True, size=11)
TITLE = Font(name=FONT, bold=True, size=14)
BAND = PatternFill("solid", fgColor="F2F5FA")
ADDED = PatternFill("solid", fgColor="E2EFDA")   # a step that did not exist
CHANGED = PatternFill("solid", fgColor="FFF2CC")  # a step whose unit moved
REST = PatternFill("solid", fgColor="FCE4D6")     # should stop being pushed


def style(ws, widths, nrows, ncols, freeze="D2"):
    for cell in ws[1]:
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(vertical="top")
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def human(step):
    dt, dv = step["durationType"], step["durationValue"]
    if dt == "TIME":
        h, m, s = dv // 3600, (dv % 3600) // 60, dv % 60
        return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"
    if dt == "DISTANCE":
        return f"{dv / 1000:g} km" if dv >= 1000 else f"{dv} m"
    return "open — lap press"


HEADERS = [
    "Day", "Week", "Day title", "Session", "Step #", "Repeat x", "Intensity",
    "Duration type", "Duration value", "Duration value type", "Reads as",
    "Target type", "Target value", "Sends today", "Change", "Why this unit",
    "Source row text",
]
WIDTHS = [6, 6, 30, 10, 7, 9, 11, 15, 15, 19, 18, 14, 13, 22, 12, 44, 62]


def main():
    plans, rebuilt_path, props_path, out = sys.argv[1:5]
    data = json.load(open(plans))
    rebuilt = json.load(open(rebuilt_path))["sessions"]
    proposals = json.load(open(props_path))["proposals"]

    by_prog = {}
    for s in rebuilt:
        by_prog.setdefault(s["programId"], []).append(s)

    programs = [p for p in data["programs"] if p["id"] in by_prog]
    programs.sort(key=lambda p: (p["programType"], p["name"], p["id"]))
    names = Counter(p["name"] for p in programs)

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Read Me")
    wb.create_sheet("Index")

    index_rows, rest_rows, change_counts = [], [], Counter()
    used = set()

    for p in programs:
        # Excel forbids [ ] : * ? / \ in a sheet name and caps it at 31 chars.
        base = re.sub(r"[\[\]:*?/\\]", "-", p["name"])
        if names[p["name"]] > 1:
            base = f"{base[:24]} ({p['id'][:6]})"
        title = base[:31].strip()
        while title.lower() in used:
            title = title[:28] + "_2"
        used.add(title.lower())

        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        meta = []

        for s in sorted(by_prog[p["id"]], key=lambda x: (x["day"], x["entryIdx"])):
            week = math.ceil(s["day"] / 7) if s["day"] > 0 else 0
            sends_today = f"{s['currentSteps']} step(s)"

            if s["kind"] == "rest":
                ws.append([s["day"], week, s["title"], "rest", None, None, "",
                           "", None, "", "nothing", "", None, sends_today,
                           "STOP", "Rest day: no distance and no duration stored.",
                           s["notes"][0] if s["notes"] else ""])
                meta.append((ws.max_row, s["day"], "rest"))
                rest_rows.append([p["name"], s["day"], s["title"], sends_today])
                change_counts["Rest day — stop pushing anything"] += 1
                continue

            for i, item in enumerate(s["steps"], start=1):
                st = item["step"]
                added = "added" in st["origin"]
                if added:
                    change = "ADDED"
                    change_counts["Warm-up or cool-down added"] += 1
                elif st["intensity"] == "RECOVERY":
                    change = "BOUNDED"
                    change_counts["Rep recovery bounded (was open)"] += 1
                else:
                    change = ""
                    if st["intensity"] in ("ACTIVE", "INTERVAL"):
                        change_counts[f"Work step sent as {st['durationType']}"] += 1

                ws.append([
                    s["day"], week, s["title"], s["kind"], i,
                    item.get("repeatOf"), st["intensity"],
                    st["durationType"], st["durationValue"], st["durationValueType"] or "",
                    human(st), st["targetType"],
                    st["targetValue"], sends_today,
                    change, st["origin"], (st["description"] or "")[:400],
                ])
                meta.append((ws.max_row, s["day"], "added" if added else
                             "changed" if st["intensity"] == "RECOVERY" else ""))

        style(ws, WIDTHS, ws.max_row - 1, len(HEADERS))

        order, seen = {}, []
        for _, dno, _ in meta:
            if dno not in order:
                order[dno] = len(seen)
                seen.append(dno)
        for r, dno, kind in meta:
            if order[dno] % 2 == 1:
                for c in range(1, len(HEADERS) + 1):
                    ws.cell(r, c).fill = BAND
            if kind == "rest":
                for c in range(1, len(HEADERS) + 1):
                    ws.cell(r, c).fill = REST
            elif kind == "added":
                for c in range(7, 12):
                    ws.cell(r, c).fill = ADDED
            elif kind == "changed":
                for c in range(7, 12):
                    ws.cell(r, c).fill = CHANGED

        sess = by_prog[p["id"]]
        index_rows.append([
            title, p["name"], p["programType"], p["id"],
            len(sess),
            sum(1 for x in sess if x["kind"] == "interval"),
            sum(1 for x in sess if x["kind"] == "steady"),
            sum(1 for x in sess if x["kind"] == "rest"),
            sum(len(x["steps"]) for x in sess),
            sum(x["currentSteps"] for x in sess),
        ])

    # ── Index ────────────────────────────────────────────────────────────────
    ws = wb["Index"]
    ws.append(["Tab", "Program", "Type", "Firestore doc ID", "Run sessions",
               "Interval-type", "Steady", "Rest (stop pushing)",
               "Steps after rebuild", "Steps today"])
    for r in index_rows:
        ws.append(r)
    style(ws, [30, 34, 9, 24, 14, 14, 10, 18, 18, 13], len(index_rows), 10, freeze="C2")

    # ── Changes ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Changes")
    ws.append(["Change", "Steps"])
    for k, n in change_counts.most_common():
        ws.append([k, n])
    style(ws, [46, 12], len(change_counts), 2, freeze="A2")

    # ── Rest Days ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Rest Days")
    ws.append(["Program", "Day", "Title", "Currently sends"])
    for r in rest_rows:
        ws.append(r)
    style(ws, [34, 7, 30, 20], len(rest_rows), 4, freeze="A2")

    # ── Strength Fixes (carried over, unchanged) ─────────────────────────────
    ws = wb.create_sheet("Strength Fixes")
    ws.append(["Program ID", "Day", "Session", "Step #", "→ Duration type",
               "→ Duration value", "→ Reads as", "→ Exercise category",
               "→ Exercise name", "Confidence", "Why"])
    strength = [x for x in proposals if x.get("durationType") in (None, "REPS", "TIME")
                and (x.get("exerciseCategory") or x.get("durationType") == "REPS"
                     or "reps not stated" in (x.get("human") or ""))]
    for x in strength:
        ws.append([x["programId"], x["day"], x["sessionIdx"], x["stepOrder"],
                   x.get("durationType") or "", x.get("durationValue"),
                   x.get("human") or "", x.get("exerciseCategory") or "",
                   x.get("exerciseName") or "", x.get("confidence") or "",
                   x.get("reason") or ""])
    style(ws, [24, 6, 9, 8, 15, 15, 34, 20, 26, 11, 96], len(strength), 11, freeze="E2")

    # ── Read Me ──────────────────────────────────────────────────────────────
    ws = wb["Read Me"]
    ws.column_dimensions["A"].width = 120
    total_steps = sum(r[8] for r in index_rows)
    lines = [
        ("Updated running plan — rebuilt for the Garmin Training API", TITLE),
        ("", BODY),
        (f"Source: {data['source']}", BODY),
        (f"Generated: {data['generatedAt']}", BODY),
        (f"{len(programs)} programs, {sum(r[4] for r in index_rows)} run sessions, "
         f"{total_steps} steps (today: {sum(r[9] for r in index_rows)}).", BODY),
        ("", BODY),
        ("NOTHING HAS BEEN WRITTEN TO FIRESTORE. This is the proposal for review.", BOLD),
        ("", BODY),
        ("How each session was rebuilt", BOLD),
        ("Every run session was rebuilt from its stored run rows rather than by regexing the joined", BODY),
        ("prose, which is what the live mapper does. The rules applied:", BODY),
        ("", BODY),
        ("Unit          One unit per step, never a mix. A duration stated in the description wins;", BODY),
        ("              then a distance stated in the description; then the stored distance.", BODY),
        ("              An advisory time is not a duration — 'a pace you could hold for approx.", BODY),
        ("              1 hour' describes an effort and 'fuel every 30-40 min' is a fuelling note.", BODY),
        ("              Both currently turn threshold runs into hour-long time steps.", BODY),
        ("              A 'time on feet' goal does count as the target, by your decision.", BODY),
        ("Warm-up       The stored warm-up row is honoured in its own unit, so a 1.5km jog stays", BODY),
        ("              DISTANCE 1500m. Only a session with none gets a 15 min TIME warm-up.", BODY),
        ("Cool-down     The same; a session with none gets an OPEN lap-press cool-down.", BODY),
        ("Recovery      Never OPEN. The recovery the row states is honoured in its own unit, so", BODY),
        ("              '400m easy jog recovery' stays DISTANCE 400m. Where none is stated: 2 min", BODY),
        ("              for reps of 1km or 3 min and over, 90 seconds for anything shorter.", BODY),
        ("Target        Every work step carries the HR zone derived from the row's own RPE, as", BODY),
        ("              targetValue 1-5 per Training API V2 section 3.2.1.", BODY),
        ("", BODY),
        ("Reading a program tab", BOLD),
        ("One row per step of the rebuilt session, in the order Garmin will receive them.", BODY),
        ("'Repeat x' is the repeat block a step sits in. 'Sends today' is what that session", BODY),
        ("currently pushes, for comparison. Green rows are steps that did not exist before;", BODY),
        ("amber rows are recoveries that were open and are now bounded; salmon rows are rest days", BODY),
        ("that should stop being pushed at all.", BODY),
        ("", BODY),
        ("Rest days", BOLD),
        ("98 rest days are stored as run rows with no distance and no duration. Because their", BODY),
        ("sessionType is 'run', the mapper skips the classifier and mapRunEasy's 30-minute default", BODY),
        ("sends each one to the watch as a half-hour easy run. They should push nothing. The Rest", BODY),
        ("Days tab lists every one.", BODY),
        ("", BODY),
        ("Strength Fixes carries the strength-side proposals unchanged from the previous workbook,", BODY),
        ("so this file is complete as an input to any write-back.", BODY),
    ]
    for i, (t, f) in enumerate(lines, start=1):
        c = ws.cell(i, 1, t)
        c.font = f
        c.alignment = Alignment(vertical="top")
    ws.sheet_view.showGridLines = False

    wb.save(out)
    print(f"Wrote {out}")
    print(f"  tabs: {wb.sheetnames}")
    print(f"  rest days flagged: {len(rest_rows)} | strength fixes carried: {len(strength)}")


if __name__ == "__main__":
    main()
