#!/usr/bin/env python3
"""Build a per-program review workbook from the training-plan export.

One tab per program, each holding the Garmin Training API payload the sync
would push, step by step, with the stored source row that produced each step
alongside it. Filtered to the program types given on the command line.

Pass a proposals file (from scripts/propose-garmin-fixes.ts) as --proposals to
add the first-pass "what this should be" columns for review.

Usage:
    python3 scripts/build-program-tabs-workbook.py plans.json out.xlsx running hybrid \
        --proposals proposals.json
"""

import json
import math
import re
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, bold=True, size=11)
TITLE = Font(name=FONT, bold=True, size=14)
BAND = PatternFill("solid", fgColor="F2F5FA")   # alternate day shading
REST_FILL = PatternFill("solid", fgColor="EDEDED")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")
MISS_FILL = PatternFill("solid", fgColor="FFF2CC")  # unresolved Garmin field
FIX_FILL = PatternFill("solid", fgColor="E2EFDA")   # a proposed change
FIX_HDR = PatternFill("solid", fgColor="375623")    # proposed-columns header

STEP_HEADERS = [
    "Day", "Week", "Day title", "Session #", "Sync key", "Session type",
    "Category", "Garmin workout name", "Garmin sport", "Est. dur (s)",
    "Step #", "In repeat", "Repeat x", "Intensity", "Step description",
    "Duration type", "Duration value", "Target type", "Target low", "Target high",
    "Exercise category", "Exercise name", "Weight", "Weight unit",
    "Source row", "Source details",
]
FIX_HEADERS = [
    "Fix?", "→ Duration type", "→ Duration value", "→ Duration value type",
    "→ Reads as", "→ Target", "→ Exercise category", "→ Exercise name",
    "Confidence", "Why",
]
FIX_WIDTHS = [6, 15, 15, 19, 24, 42, 20, 26, 11, 96]
STEP_WIDTHS = [6, 6, 30, 9, 9, 12, 14, 30, 20, 11,
               7, 10, 9, 11, 46, 14, 14, 12, 11, 11,
               19, 26, 8, 11, 26, 52]
NSTEP = len(STEP_HEADERS)


# Rows that read as coaching prose rather than work. They become steps the
# athlete has to lap through on the watch.
# Matches workout-mapper.ts's NOTE_ROW exactly — "format" and "work N" are
# deliberately excluded there (they carry the circuit's actual round
# structure and prescribed movements, not commentary), so they must be
# excluded here too or this flag would fire on rows the mapper is right to
# keep as steps.
NOTE_ROW = re.compile(r"^(notes?|session\s+notes|coaching\s+notes?|focus|cue|reminder)\b", re.I)

# The circuit-header rows the mapper deliberately keeps as steps (they carry
# real content), but which were never going to resolve to a Garmin exercise.
FORMAT_OR_WORK_ROW = re.compile(r"^(format|work\s*\d*)\b", re.I)

# A rest day stored as an empty run row. Pushing nothing for one is correct,
# so it must not be reported as a day that reaches the watch with nothing.
REST_TITLE = re.compile(r"^(rest|rest or cross[- ]?train|what's next|full rest)", re.I)

# A loose "looks like it states sets and reps somewhere" check, used only to
# decide whether MAPPER_SETS_REPS (below) is worth running at all. Carries the
# same distance/timed exclusion so "3 x 20m" (a loaded-carry distance) isn't
# treated as a candidate rep count in the first place. The \b before the
# lookahead matters: without it, "20m" lets (\d+) backtrack to "2" \u2014 "0" and
# "m" are both word characters, so the lookahead no longer sees "m" right
# there and wrongly reports a match on "20m" as if it read "2".
SETS_REPS = re.compile(
    r"(\d+)\s*(?:sets?)?\s*[x\u00d7]\s*(\d+)\b(?!\s*(?:m\b|sec|s\b))"
    r"|(\d+)\s*sets?\s+of\s+(\d+)\b(?!\s*(?:m\b|sec|s\b))",
    re.I,
)

# Mirrors parseSetsReps() in workout-mapper.ts exactly, including the
# negative lookahead that excludes a distance or timed shape ("3 x 20m",
# "3x90sec") so a loaded-carry distance is never misread as a rep count.
# Kept in sync by hand \u2014 if that function changes, update this too.
MAPPER_SETS_REPS = re.compile(
    r"(\d+)\s*(?:sets?)?\s*[x\u00d7]\s*(\d+)(?:\s*[-\u2013]\s*(\d+))?\b(?!\s*(?:m\b|sec|s\b))"
    r"|(\d+)\s*sets?\s+of\s+(\d+)\b(?!\s*(?:m\b|sec|s\b))",
    re.I,
)


def flatten(steps, repeat_no=None, repeat_val=None):
    """Yield (repeat_no, repeat_value, step), descending into repeat blocks."""
    for st in steps:
        if st.get("type") == "WorkoutRepeatStep":
            yield from flatten(st.get("steps", []), st.get("stepOrder"), st.get("repeatValue"))
        else:
            yield (repeat_no, repeat_val, st)


def all_steps(workout):
    if not workout:
        return []
    return [t for seg in workout["segments"] for t in flatten(seg.get("steps", []))]


def match_source(desc, day):
    """Find the stored row a step came from.

    mapStrength writes '<exercise name> — ...', so an exact prefix match is
    reliable. Run sessions produce one step from one planned run, so a lone
    run attributes unambiguously. Anything else is left blank rather than
    guessed at.
    """
    if not desc:
        return ("", "")
    for ex in day["exercises"]:
        name = (ex.get("name") or "").strip()
        if name and (desc.startswith(name + " —") or desc.startswith(name + ":") or desc == name):
            return (name, ex.get("details") or "")
    if len(day["runs"]) == 1 and not day["exercises"]:
        r = day["runs"][0]
        return (f"run: {r.get('type', '')}", r.get("description") or "")
    return ("", "")


def sheet_title(program, used, ambiguous):
    """Excel tab name: <=31 chars, no []:*?/\\, unique across the workbook.

    Programs whose names collide are all suffixed with a slice of the doc ID,
    not just the second one — two tabs called the same thing with only one
    labelled is worse than neither being labelled.
    """
    base = re.sub(r"[\[\]:*?/\\]", "-", program["name"]).strip()
    if program["name"] in ambiguous:
        suffix = f" ({program['id'][:6]})"
        name = base[: 31 - len(suffix)].strip() + suffix
    else:
        name = base[:31].strip()
    n = 2
    while name.lower() in used:
        suffix = f" ({n})"
        name = name[: 31 - len(suffix)].strip() + suffix
        n += 1
    used.add(name.lower())
    return name


def style_table(ws, headers, widths, nrows, freeze="D2"):
    for c, cell in enumerate(ws[1], start=1):
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(vertical="top")
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{nrows + 1}"


def build_program_sheet(wb, p, used_titles, ambiguous, props):
    ws = wb.create_sheet(sheet_title(p, used_titles, ambiguous))
    ws.append(STEP_HEADERS + (FIX_HEADERS if props is not None else []))

    rows_meta = []   # (row_index, day_number, is_rest, missing_category)
    for day in p["days"]:
        dno, title = day["day"], day["title"]
        week = math.ceil(dno / 7) if dno > 0 else 0

        for s in day["sessions"]:
            w = s["workout"]
            steps = all_steps(w)

            # A session the mapper declined still gets a row, otherwise a day
            # that silently reaches nothing on the watch just isn't in the sheet.
            if not steps:
                ws.append([dno, week, title, s["sessionIdx"], s["dayKey"],
                           s.get("sessionType") or "", s["category"],
                           "— nothing pushed to Garmin —", "", None,
                           None, "", None, "", "", "", None, "", None, None,
                           "", "", None, "", "", ""])
                rows_meta.append((ws.max_row, dno, True, False, False))
                continue

            for repeat_no, repeat_val, st in steps:
                src_name, src_details = match_source(st.get("description"), day)
                fx = props.get((p["id"], day["entryIdx"], s["sessionIdx"], st["stepOrder"])) \
                    if props is not None else None
                ws.append([
                    dno, week, title, s["sessionIdx"], s["dayKey"],
                    s.get("sessionType") or "", s["category"],
                    w["workoutName"], w["sport"], w.get("estimatedDurationInSecs"),
                    st["stepOrder"], repeat_no or "", repeat_val or None,
                    st["intensity"], st.get("description") or "",
                    st["durationType"], st.get("durationValue"),
                    st["targetType"], st.get("targetValueLow"), st.get("targetValueHigh"),
                    st.get("exerciseCategory") or "", st.get("exerciseName") or "",
                    st.get("weightValue"), st.get("weightDisplayUnit") or "",
                    src_name, src_details,
                ] + ([
                    "FIX" if fx else "", fx.get("durationType") or "" if fx else "",
                    fx.get("durationValue") if fx else None,
                    fx.get("durationValueType") or "" if fx else "",
                    fx.get("human") or "" if fx else "", fx.get("target") or "" if fx else "",
                    fx.get("exerciseCategory") or "" if fx else "",
                    fx.get("exerciseName") or "" if fx else "",
                    fx.get("confidence") or "" if fx else "", fx.get("reason") or "" if fx else "",
                ] if props is not None else []))
                missing = (st["intensity"] == "ACTIVE"
                           and w["sport"] == "STRENGTH_TRAINING"
                           and not st.get("exerciseCategory"))
                rows_meta.append((ws.max_row, dno, False, missing, bool(fx)))

        if not day["sessions"]:
            ws.append([dno, week, title, "", "", "", "(no session generated)",
                       "— nothing pushed to Garmin —", "", None,
                       None, "", None, "", "", "", None, "", None, None,
                       "", "", None, "", "", ""])
            rows_meta.append((ws.max_row, dno, True, False, False))

    style_table(ws, STEP_HEADERS + (FIX_HEADERS if props is not None else []),
                STEP_WIDTHS + (FIX_WIDTHS if props is not None else []), ws.max_row - 1)
    if props is not None:
        for c in range(NSTEP + 1, NSTEP + len(FIX_HEADERS) + 1):
            ws.cell(1, c).fill = FIX_HDR

    # Shade alternate day numbers so each day reads as a block, and highlight
    # the two things a reviewer is scanning for.
    order, seen = {}, []
    for _, dno, _, _, _ in rows_meta:
        if dno not in order:
            order[dno] = len(seen)
            seen.append(dno)
    ncols = NSTEP + (len(FIX_HEADERS) if props is not None else 0)
    for r, dno, is_rest, missing, has_fix in rows_meta:
        if order[dno] % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = BAND
        if is_rest:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = REST_FILL
        if missing:
            ws.cell(r, 21).fill = MISS_FILL
        if has_fix:
            for c in range(NSTEP + 1, ncols + 1):
                ws.cell(r, c).fill = FIX_FILL
    return ws


def program_flags(p):
    """Review findings for one program."""
    out = []
    pid, pname = p["id"], p["name"]
    day_counts = Counter(d["day"] for d in p["days"])
    key_counts = Counter(s["dayKey"] for d in p["days"] for s in d["sessions"])

    reported = set()
    for d in p["days"]:
        dno, title = d["day"], d["title"]
        sessions, mapped = d["sessions"], [s for s in d["sessions"] if s["workout"]]

        if day_counts[dno] > 1 and dno not in reported:
            reported.add(dno)
            colliding = [k for k in {s["dayKey"] for x in p["days"] if x["day"] == dno
                                     for s in x["sessions"]} if key_counts[k] > 1]
            if colliding:
                # Should not happen post-fix (sync-keys.ts gives every colliding
                # entry a distinct key) — kept as a canary in case that ever
                # regresses, rather than assuming it can't.
                out.append([pname, dno, title, "Sync key collision",
                            f"Day {dno} is stored as {day_counts[dno]} separate workouts[] entries, and "
                            f"{len(colliding)} of their sync keys collide ({', '.join(sorted(colliding))}). "
                            "Both are created on Garmin, but garminPlanSync records only the last, so the "
                            "earlier one is never cleaned up on the next sync."])
            else:
                out.append([pname, dno, title, "Paired sessions share a day number",
                            f"Day {dno} is stored as {day_counts[dno]} separate workouts[] entries (e.g. a "
                            "run and a strength day). This is fine — sync-keys.ts gives each one its own "
                            "sync key so they don't overwrite each other — but worth a glance to confirm "
                            "it's the intended pairing, not an accidental duplicate."])

        is_rest = bool(REST_TITLE.match(title.strip())) and not any(
            (r.get("distance") or 0) > 0 for r in d["runs"])
        if (d["exercises"] or d["runs"]) and not mapped and not is_rest:
            out.append([pname, dno, title, "Nothing pushed to Garmin",
                        f"Day has content but produced no Garmin workout "
                        f"(category: {', '.join(sorted({s['category'] for s in sessions})) or 'none'})."])
        for msg in d.get("dataFlags", []):
            out.append([pname, dno, title, "Stored data shape", msg])

        # A canary, not an expected finding: workout-mapper.ts strips a
        # NOTE_ROW-matching exercise from the step list (its text still
        # reaches the workout description, just not as a step), so this
        # checks the actual emitted steps rather than the stored exercise
        # list \u2014 the source row is untouched and would always match.
        note_names = {(e.get("name") or "").strip() for e in d["exercises"]
                      if NOTE_ROW.match((e.get("name") or "").strip()
                                        .lstrip("\U0001F4CB\U0001F511\U0001F4A1\u26A0\uFE0F ").strip())}
        if note_names:
            emitted = {(st.get("description") or "").split(" \u2014 ")[0].split(":")[0].strip()
                       for s in mapped for _, _, st in all_steps(s["workout"])}
            leaked = note_names & emitted
            if leaked:
                out.append([pname, dno, title, "Prose row sent as a step",
                            f"{', '.join(sorted(leaked))} matches the coaching-prose pattern but still "
                            "appears as a step description \u2014 check NOTE_ROW in workout-mapper.ts."])

        for s in mapped:
            w = s["workout"]
            steps = all_steps(w)
            actives = [st for _, _, st in steps if st["intensity"] == "ACTIVE"]

            if w["sport"] == "STRENGTH_TRAINING":
                # "Format"/"Work N" rows are prescribed content, not an
                # exercise — they were never going to have a Garmin category,
                # so counting them here would misreport a design choice as a
                # lookup-table gap.
                unmatched = [st for st in actives if not st.get("exerciseCategory")
                             and not FORMAT_OR_WORK_ROW.match((st.get("description") or "").split(":")[0].strip())]
                if unmatched:
                    names = sorted({(st.get("description") or "").split(" — ")[0]
                                    for st in unmatched})[:6]
                    out.append([pname, dno, title, "No Garmin exercise match",
                                f"{len(unmatched)} of {len(actives)} work steps carry no exerciseCategory, so "
                                f"they appear unnamed on the watch: {', '.join(names)}."])

            # A strength exercise whose sets/reps never parsed becomes one
            # open-ended step: no rep count, no set structure, nothing to
            # auto-advance. Aggregated per session to stay readable.
            if w["sport"] == "STRENGTH_TRAINING":
                unparsed = []
                for repeat_no, _, st in steps:
                    if repeat_no is not None or st["intensity"] != "ACTIVE":
                        continue
                    if st["durationType"] != "OPEN":
                        continue
                    desc = st.get("description") or ""
                    nm = desc.split(" — ")[0].split(":")[0].strip()
                    src = next((e for e in d["exercises"] if (e.get("name") or "").strip() == nm), None)
                    det = (src or {}).get("details") or ""
                    if SETS_REPS.search(det) and not MAPPER_SETS_REPS.search(det):
                        unparsed.append((nm, det))
                if unparsed:
                    out.append([pname, dno, title, "Sets/reps not parsed",
                                f"{len(unparsed)} exercise(s) state sets and reps that parseSetsReps() does "
                                f"not match (it accepts an ASCII 'x' only), so each maps to a single "
                                f"open-ended step with no rep target: "
                                + "; ".join(f"{n} ({t[:40]})" for n, t in unparsed[:4]) + "."])

            if actives and all(st["durationType"] == "OPEN" for st in actives):
                out.append([pname, dno, title, "No measurable targets",
                            f"All {len(actives)} work steps are OPEN duration with no target — the watch "
                            "cannot auto-advance or measure this session."])

            if w["sport"] == "RUNNING" and d["runs"] and s["category"] != "run_intervals":
                km = sum(r.get("distance") or 0 for r in d["runs"])
                secs = w.get("estimatedDurationInSecs") or 0
                if len(steps) == 1 and secs == 1800 and km > 0:
                    out.append([pname, dno, title, "Run duration defaulted",
                                f"No duration in the text, so the step defaulted to 30:00 — the planned "
                                f"{km}km is never used."])
                elif len(steps) == 1 and km > 0 and secs > 0:
                    pace = secs / km
                    if pace < 150 or pace > 900:
                        out.append([pname, dno, title, "Duration / distance mismatch",
                                    f"Step is {secs // 60}:{secs % 60:02d} for a planned {km}km "
                                    f"({pace / 60:.1f} min/km). Duration was read from the description "
                                    "text, not the distance."])
                # Only a genuine flattening: the session emitted no repeat block.
                has_repeat = any(
                    x.get("type") == "WorkoutRepeatStep"
                    for seg in w["segments"] for x in seg.get("steps", []))
                # A row typed 'intervals' with a single rep is a continuous
                # effort — a time trial — not a flattened interval session.
                iv = [r for r in d["runs"] if (r.get("noIntervals") or 0) > 1]
                if iv and not has_repeat:
                    out.append([pname, dno, title, "Intervals flattened",
                                f"{len(iv)} interval/tempo run(s) stored, but the day classified as "
                                f"'{s['category']}', so it maps to one steady run. The interval structure "
                                "never reaches the watch."])
    return out


def main():
    argv = sys.argv[1:]
    prop_path = None
    if "--proposals" in argv:
        i = argv.index("--proposals")
        prop_path = argv[i + 1]
        argv = argv[:i] + argv[i + 2:]
    src, out = argv[0], argv[1]
    wanted = [a.lower() for a in argv[2:]] or ["running", "hybrid"]

    props = restructures = None
    if prop_path:
        raw = json.load(open(prop_path))
        props = {(x["programId"], x["entryIdx"], x["sessionIdx"], x["stepOrder"]): x
                 for x in raw["proposals"]}
        restructures = raw["restructures"]

    data = json.load(open(src))
    programs = [p for p in data["programs"] if (p["programType"] or "").lower() in wanted]
    programs.sort(key=lambda p: (p["programType"], p["name"], p["id"]))

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Read Me")
    wb.create_sheet("Index")

    name_counts = Counter(p["name"] for p in programs)
    ambiguous = {n for n, c in name_counts.items() if c > 1}

    used, index_rows, flag_rows, source_rows = set(), [], [], []
    for p in programs:
        ws = build_program_sheet(wb, p, used, ambiguous, props)
        steps_n = sum(len(all_steps(s["workout"])) for d in p["days"] for s in d["sessions"])
        sess_n = sum(len(d["sessions"]) for d in p["days"])
        pushed = sum(1 for d in p["days"] for s in d["sessions"] if s["workout"])
        index_rows.append([
            ws.title, p["name"], p["programType"], p.get("targetRace") or "",
            p["id"], p["collection"], p.get("visibility") or "public",
            p["assignedUserCount"], len(p["days"]),
            len({d["day"] for d in p["days"]}),
            math.ceil(max((d["day"] for d in p["days"]), default=0) / 7),
            sess_n, pushed, sess_n - pushed, steps_n,
        ])
        flag_rows.extend(program_flags(p))

        for d in p["days"]:
            for i, ex in enumerate(d["exercises"], 1):
                source_rows.append([p["name"], p["id"], d["day"], d["title"], "exercise", i,
                                    ex.get("name", ""), ex.get("details", ""),
                                    ex.get("sessionType", ""), ex.get("garminSport", ""),
                                    ex.get("garminExerciseCategory", ""), ex.get("garminExerciseName", ""),
                                    ex.get("sets"), ex.get("reps"), ex.get("weightKg"),
                                    ex.get("restSeconds"), None, "", None, None, None])
            for i, r in enumerate(d["runs"], 1):
                source_rows.append([p["name"], p["id"], d["day"], d["title"], "run", i,
                                    r.get("type", ""), r.get("description", ""),
                                    "run", "RUNNING", "", "", None, None, None, None,
                                    r.get("distance"), r.get("paceZone", ""), r.get("targetPace"),
                                    r.get("effortLevel"), r.get("noIntervals")])

    # ── Index ────────────────────────────────────────────────────────────────
    ws = wb["Index"]
    ws.append(["Tab", "Program", "Type", "Target race", "Firestore doc ID", "Collection",
               "Visibility", "Assigned users", "Stored days", "Distinct day numbers",
               "Weeks", "Sessions", "Pushed to Garmin", "Not pushed", "Garmin steps"])
    for r in index_rows:
        ws.append(r)
    style_table(ws, [None] * 15,
                [26, 34, 9, 12, 24, 13, 11, 14, 12, 20, 7, 10, 16, 12, 13],
                len(index_rows), freeze="C2")
    for r in range(2, len(index_rows) + 2):
        # Live count of the rows actually written on that program's tab.
        tab = ws.cell(r, 1).value
        ws.cell(r, 15, f"=COUNTA('{tab}'!$A$2:$A$100000)").font = BODY
    note = ws.cell(len(index_rows) + 3, 1,
                   "Counts are extracted values except 'Garmin steps', which counts the rows on "
                   "each program tab. Rows for sessions that push nothing are included there.")
    note.font = Font(name=FONT, size=9, italic=True, color="595959")

    # ── Proposed Fixes (summary) ─────────────────────────────────────────────
    if props is not None:
        by_prog = Counter()
        by_kind = Counter()
        for x in props.values():
            pname = next((r[1] for r in index_rows if r[4] == x["programId"]), x["programId"])
            dt, dc = x.get("durationType"), x.get("durationChanged")
            kind = ("Duration: use the stored distance" if dt == "DISTANCE" and dc
                    else "Duration: re-read the stated time" if dt == "TIME" and dc
                    else "Sets/reps: add a rep target" if dt == "REPS"
                    else "Sets stated, reps missing from the plan" if "reps not stated" in (x.get("human") or "")
                    else "Exercise: add the Garmin category" if x.get("exerciseCategory")
                    else "Target: add an HR zone from RPE")
            by_prog[(pname, kind)] += 1
            by_kind[kind] += 1

        ws = wb.create_sheet("Proposed Fixes")
        ws.append(["Change", "Steps affected", "Programs"])
        for kind, n in by_kind.most_common():
            progs_hit = sorted({k[0] for k in by_prog if k[1] == kind})
            ws.append([kind, n, ", ".join(progs_hit)])
        style_table(ws, [None] * 3, [40, 15, 96], len(by_kind), freeze="A2")
        r0 = len(by_kind) + 3
        ws.cell(r0, 1, "Per program").font = BOLD
        ws.cell(r0 + 1, 1, "Program").font = HDR_FONT
        ws.cell(r0 + 1, 1).fill = HDR_FILL
        ws.cell(r0 + 1, 2, "Change").font = HDR_FONT
        ws.cell(r0 + 1, 2).fill = HDR_FILL
        ws.cell(r0 + 1, 3, "Steps").font = HDR_FONT
        ws.cell(r0 + 1, 3).fill = HDR_FILL
        for i, ((pname, kind), n) in enumerate(sorted(by_prog.items()), start=r0 + 2):
            ws.cell(i, 1, pname).font = BODY
            ws.cell(i, 2, kind).font = BODY
            ws.cell(i, 3, n).font = BODY

    # ── Spec Conformance ─────────────────────────────────────────────────────
    # Checked against Garmin Connect Developer Program, Training API V2,
    # version 1.0 (26/05/2025), section 3.2.1 Field Definitions.
    ws = wb.create_sheet("Spec Conformance")
    ws.append(["Field / rule", "Training API V2 says", "What we send", "Verdict"])
    conformance = [
        ("sport",
         "RUNNING, CYCLING, LAP_SWIMMING, STRENGTH_TRAINING, CARDIO_TRAINING, GENERIC, YOGA, PILATES",
         "RUNNING, STRENGTH_TRAINING, CARDIO_TRAINING", "OK"),
        ("intensity",
         "REST, WARMUP, COOLDOWN, RECOVERY, ACTIVE, INTERVAL, MAIN (swim only)",
         "WARMUP, ACTIVE, INTERVAL, RECOVERY, REST, COOLDOWN", "OK"),
        ("durationType",
         "TIME, DISTANCE, HR_LESS_THAN, HR_GREATER_THAN, CALORIES, OPEN, POWER_*, FIXED_REST; "
         "REPS for HIIT / CARDIO / STRENGTH_TRAINING only",
         "TIME, DISTANCE, OPEN, REPS — REPS only on STRENGTH_TRAINING", "OK"),
        ("repeatType",
         "REPEAT_UNTIL_STEPS_CMPLT and seven conditional variants",
         "REPEAT_UNTIL_STEPS_CMPLT", "OK"),
        ("Steps per workout",
         "100 for a single-segment workout",
         "17 at most", "OK"),
        ("Step description",
         "512 characters, longer is truncated",
         "206 at most", "OK"),
        ("workoutProvider / workoutSourceId",
         "20 characters max", '"HybridX"', "OK"),
        ("weightValue / weightDisplayUnit",
         "Kilograms; display unit KILOGRAM or POUND; STRENGTH_TRAINING only",
         "Always null — no program stores weightKg", "Gap: no weights sent"),
        ("targetType",
         "SPEED, HEART_RATE, CADENCE, POWER, GRADE, RESISTANCE, POWER_*, SPEED_LAP, "
         "HEART_RATE_LAP, OPEN, PACE (as speed in m/s)",
         "OPEN on all 1,672 steps", "Gap: no targets sent"),
        ("targetValue",
         "The target HR zone (1-5) or power zone (1-7). Zones must already be defined "
         "and saved on the athlete's Garmin account.",
         "buildStep() hardcodes targetValue: null and exposes only "
         "targetValueLow/High, so a zone target cannot be expressed at all",
         "Gap: blocks the HR-zone fix"),
        ("targetValueLow / High",
         "A custom range, used instead of a zone in targetValue",
         "Set only by paceTarget(), which never fires because no program stores targetPace",
         "Unused in practice"),
        ("durationValueType",
         "Field description says it modifies HR and POWER only (PERCENT), but every "
         "DISTANCE example in the spec sets it to METER",
         "Always null, including on the 75 DISTANCE steps",
         "Confirm with Garmin"),
        ("estimatedDurationInSecs",
         "Calculated server-side; ignored on Create and Update",
         "Sent on run workouts",
         "Harmless, but it is not what drives the watch — the step's own "
         "durationValue is"),
        ("exerciseCategory / exerciseName",
         "STRENGTH_TRAINING, CARDIO_TRAINING, HIIT, PILATES, YOGA only. Valid values are "
         "in Appendix A and B, supplied as a separate Excel file.",
         "Values come from the lookup table in program-enricher.ts",
         "Unverified: Appendix B was not in the PDF"),
    ]
    for r in conformance:
        ws.append(list(r))
    style_table(ws, [None] * 4, [30, 74, 62, 34], len(conformance), freeze="A2")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value and cell.value != "OK":
                cell.fill = FLAG_FILL

    # ── Session Restructures ─────────────────────────────────────────────────
    if restructures:
        ws = wb.create_sheet("Session Restructures")
        ws.append(["Program", "Day", "Day title", "Currently", "Proposed structure", "Why"])
        for x in restructures:
            ws.append([x["programName"], x["day"], x["title"], x["current"],
                       x["proposed"], x["reason"]])
        style_table(ws, [None] * 6, [30, 6, 28, 24, 74, 92], len(restructures), freeze="D2")

    # ── Review Flags ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Review Flags")
    ws.append(["Program", "Day", "Day title", "Flag", "Detail"])
    for r in flag_rows:
        ws.append(r)
    style_table(ws, [None] * 5, [34, 6, 32, 26, 104], len(flag_rows), freeze="A2")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.fill = FLAG_FILL

    # ── All Source Rows ──────────────────────────────────────────────────────
    ws = wb.create_sheet("All Source Rows")
    ws.append(["Program", "Doc ID", "Day", "Day title", "Kind", "#", "Name / run type",
               "Details / description", "Session type", "Garmin sport",
               "Garmin exercise category", "Garmin exercise name", "Sets", "Reps",
               "Weight (kg)", "Rest (s)", "Distance (km)", "Pace zone",
               "Target pace (s/km)", "Effort (RPE)", "Intervals"])
    for r in source_rows:
        ws.append(r)
    style_table(ws, [None] * 21,
                [30, 22, 6, 30, 9, 4, 28, 64, 12, 19, 23, 25, 6, 6, 11, 8, 12, 11, 17, 11, 9],
                len(source_rows), freeze="D2")

    # ── Read Me ──────────────────────────────────────────────────────────────
    ws = wb["Read Me"]
    ws.column_dimensions["A"].width = 122
    lines = [
        ("Running & hybrid training plans — Garmin sync review", TITLE),
        ("", BODY),
        (f"Source: {data['source']}", BODY),
        (f"Generated: {data['generatedAt']}", BODY),
        (f"Programs: {len(programs)} ({', '.join(sorted(set(p['programType'] for p in programs)))})", BODY),
        ("", BODY),
        ("How this was produced", BOLD),
        ("Every stored program day was run through the same code the live sync uses:", BODY),
        ("    program.workouts[]  →  workoutToDays()  →  mapWorkoutDay()  →  Garmin workout", BODY),
        ("imported directly from src/lib/garmin/program-adapter.ts and workout-mapper.ts. The payload", BODY),
        ("columns are therefore what POST /api/garmin/sync-plan would send to the Garmin Training API,", BODY),
        ("not a separate reading of the mapper.", BODY),
        ("", BODY),
        ("Reading a program tab", BOLD),
        ("One row per Garmin workout step. Columns A–J identify the day and the workout it belongs to;", BODY),
        ("K–X are the Training API step payload; Y–Z are the stored row that produced the step.", BODY),
        ("Shading alternates per day number so each day reads as a block. Grey rows are days that reach", BODY),
        ("the watch with nothing. An amber 'Exercise category' cell is a strength step Garmin cannot name.", BODY),
        ("", BODY),
        ("Column notes", BOLD),
        ("Sync key         The key the sync writes into users/{uid}.garminPlanSync.workouts.", BODY),
        ("Category         What classifyWorkout() made of the session. For run sessions it only decides", BODY),
        ("                 intervals vs steady; for days with no explicit sessionType it decides everything.", BODY),
        ("In repeat        Step order of the WorkoutRepeatStep this step sits inside; blank if top level.", BODY),
        ("Repeat x         How many times that block runs.", BODY),
        ("Source row       Matched back by name — exact for strength steps, and for run days with a single", BODY),
        ("                 planned run. Blank where the step has no unambiguous source.", BODY),
        ("", BODY),
        ("Assumptions and limits", BOLD),
        ("• Weight is blank throughout because no exercise in these programs stores weightKg (0 of 592).", BODY),
        ("  At sync time resolveWeightKg() can still derive a load from the athlete's 1RM personal records,", BODY),
        ("  so a real push may carry weights this extract cannot show — that part is per-athlete.", BODY),
        ("• Exercise category is stored on 135 of 592 exercises (23%). For the rest, the value shown was", BODY),
        ("  resolved at sync time by lookupGarminExercise() matching against the exercise name.", BODY),
        ("• sets, reps and restSeconds are stored on none of the 592 exercises. enrichExerciseWithGarmin()", BODY),
        ("  exists to precompute them, but has not been run over these programs, so every sync re-parses", BODY),
        ("  them out of the free-text details with parseSetsReps(). That is why its gaps matter — see the", BODY),
        ("  'Sets/reps not parsed' flag.", BODY),
        ("• The sync's ~14-day horizon is ignored: every day of every program is mapped so the whole plan", BODY),
        ("  can be reviewed at once.", BODY),
        ("", BODY),
        ("The proposed-fix columns (green header)", BOLD),
        ("A first pass at what each step should carry, derived from the stored program data rather", BODY),
        ("than from the mapper's current output. NOTHING HAS BEEN CHANGED — these are for your review.", BODY),
        ("'Fix?' marks a row with a proposal. The '→' columns hold the proposed value, and 'Why' says", BODY),
        ("what it was read from. Three kinds of change:", BODY),
        ("  Duration, re-read from text   The session states its own length ('1 hr 45 min', '2 hrs').", BODY),
        ("                                parseDurationMinutes() takes the first 'N min' it sees, so", BODY),
        ("                                '1 hr 45 min' currently becomes 45 minutes.", BODY),
        ("  Duration from stored distance The running plans hold no duration in the text but do store", BODY),
        ("                                distance in km. Those are distance runs and should be sent as", BODY),
        ("                                DISTANCE in metres, not a defaulted 30:00.", BODY),
        ("  Target from RPE               Every one of the 1,672 steps ships targetType OPEN, because no", BODY),
        ("                                program sets targetPace. effortLevel is stored on all 811 runs", BODY),
        ("                                and rpeToHrZone() already maps it to a zone — it is computed at", BODY),
        ("                                sync time and then never used.", BODY),
        ("", BODY),
        ("Target shape is now settled against the spec (Training API V2 v1.0, 26/05/2025, §3.2.1): a zone", BODY),
        ("target sets targetType HEART_RATE and puts the zone 1-5 in targetValue, leaving targetValueLow", BODY),
        ("and targetValueHigh null — those two are for a custom bpm range instead of a zone. Note that", BODY),
        ("buildStep() hardcodes targetValue to null and never exposes it, so this fix needs that opened up", BODY),
        ("first, and the athlete must already have HR zones saved on their Garmin account.", BODY),
        ("", BODY),
        ("See the Spec Conformance tab for the field-by-field check, including the one open question", BODY),
        ("(durationValueType on DISTANCE steps) and the one thing the PDF could not settle (the exercise", BODY),
        ("category enum lives in Appendix B, a separate Excel file).", BODY),
        ("", BODY),
        ("Session Restructures lists days whose whole step structure is wrong, not just a value —", BODY),
        ("stored as warm-up / reps / recovery / cool-down rows but flattened into one step.", BODY),
        ("", BODY),
        ("What the flags mean", BOLD),
        ("Sync key collision         Two workouts[] entries share a day number AND produce the same sync key —", BODY),
        ("                           sync-keys.ts should make this impossible; a report here is a regression,", BODY),
        ("                           not an expected finding.", BODY),
        ("Paired sessions share a    Two workouts[] entries share a day number (e.g. a run + a strength day).", BODY),
        ("day number                 Each gets its own sync key, so this is informational — confirm it's the", BODY),
        ("                           intended pairing, not an accidental duplicate.", BODY),
        ("No Garmin exercise match   Strength steps with no exerciseCategory — unnamed generic steps on the", BODY),
        ("                           watch. Fix by renaming so program-enricher.ts matches, or setting it.", BODY),
        ("No measurable targets      Every work step is OPEN with no target; the watch cannot auto-advance.", BODY),
        ("Sets/reps not parsed       The exercise states sets and reps, but in a shape parseSetsReps() does", BODY),
        ("                           not match — it accepts an ASCII 'x' only, so '3 sets x 20 reps' with a", BODY),
        ("                           Unicode multiplication sign falls through and the step loses its", BODY),
        ("                           structure entirely.", BODY),
        ("Run duration defaulted     mapRunEasy() found no 'N minutes' in the text, so the step is a flat", BODY),
        ("                           30:00 and the planned distance is ignored.", BODY),
        ("Duration / distance        The step duration implies an impossible pace for the planned distance,", BODY),
        ("mismatch                   because it was read out of the description text.", BODY),
        ("Intervals flattened        Interval or tempo runs on a day the classifier did not read as intervals.", BODY),
        ("Nothing pushed to Garmin   The day holds content but the mapper returned null for every session.", BODY),
        ("Prose row sent as a step   Canary, not an expected finding: a 'Notes'/'Coaching Note' row still", BODY),
        ("                           reached the watch as a step. workout-mapper.ts's NOTE_ROW should", BODY),
        ("                           already exclude these — 'Format' and 'Work N' rows are deliberately", BODY),
        ("                           NOT excluded, since they carry the circuit's actual round structure", BODY),
        ("                           and prescribed movements, not commentary.", BODY),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(i, 1, text)
        c.font = font
        c.alignment = Alignment(vertical="top")
    ws.sheet_view.showGridLines = False

    wb.save(out)
    print(f"Wrote {out}")
    print(f"  tabs: {wb.sheetnames}")
    print(f"  programs {len(programs)} | flags {len(flag_rows)} | source rows {len(source_rows)}")


if __name__ == "__main__":
    main()
